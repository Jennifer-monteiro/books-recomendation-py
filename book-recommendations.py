import openpyxl
import random
filename = "books.xlsx"

workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

workbook.close()


def search_book():
    books = read_book_file("books.xlsx")
    if not books:
        print("No books in the database")
        return

    
    book_choice = input("Choose a book by genre, or author: ").lower()
    matching_list = []

    for book in books:
        if book_choice in book["genre"].lower() or book_choice in book["author"]:
            matching_list.append(book)

    if matching_list:
        result = random.choice(matching_list)
        return result
    else:
        print("No matching books found.")
        return None


def read_book_file(books_file):
    filename = books_file
    
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    books_list = []
    for row in sheet.iter_rows(values_only=True):
        book = {
            "title": row[0],
            "author": row[1],
            "genre": row[2],
            "description": row[3]
        }
        books_list.append(book)
    workbook.close()
    return books_list


def printing_info(result):
    print(f"Title: {result['title']}")
    print(f"Author: {result['author']}")
    print(f"Genre: {result['genre']}")
    print(f"Description: {result['description']}")
    print()

result = search_book()
printing_info(result)
